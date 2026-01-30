from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from jose import jwt, JWTError
from twilio.rest import Client
import openpyxl
import io

# =====================
# ENV SETUP
# =====================

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

TWILIO_ACCOUNT_SID = os.environ.get("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_VERIFY_SID = os.environ.get("TWILIO_VERIFY_SID", "")
ADMIN_PHONE = os.environ.get("ADMIN_PHONE", "")

ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "12345")

SECRET_KEY = os.environ.get("SESSION_SECRET", "change-me")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60

# =====================
# APP SETUP
# =====================

app = FastAPI()
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

pending_otp_sessions = {}

# =====================
# MODELS
# =====================

class Subject(BaseModel):
    semester: str
    code: str
    grade: str
    status: Optional[str] = None

class StudentCreate(BaseModel):
    rollNo: str
    name: str
    dob: str
    course: str
    subjects: List[Subject]

class LoginRequest(BaseModel):
    username: str
    password: str

class OTPVerifyRequest(BaseModel):
    code: str

# =====================
# HELPERS
# =====================

def get_result_status(grade: str) -> str:
    pass_grades = ["O", "A+", "A", "B+", "B", "C"]
    return "Pass" if grade.upper() in pass_grades else "Fail"

def create_access_token(data: dict, expires: int = 15):
    payload = data.copy()
    payload["exp"] = datetime.now(timezone.utc) + timedelta(minutes=expires)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(token: str):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        return None

async def get_current_admin(token: str):
    payload = verify_token(token)
    if not payload or not payload.get("is_admin"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    return payload

# =====================
# ROUTES
# =====================

@api_router.get("/")
async def root():
    return {"message": "College Result Portal API"}

@api_router.post("/login")
async def admin_login(req: LoginRequest):
    if req.username != ADMIN_USER or req.password != ADMIN_PASS:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    if not all([TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_VERIFY_SID, ADMIN_PHONE]):
        token = create_access_token(
            {"sub": req.username, "is_admin": True},
            ACCESS_TOKEN_EXPIRE_MINUTES
        )
        return {"otp_required": False, "token": token}

    try:
        twilio = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        twilio.verify.v2.services(TWILIO_VERIFY_SID).verifications.create(
            to=ADMIN_PHONE, channel="sms"
        )

        session_id = create_access_token({"pending_otp": True}, 10)
        pending_otp_sessions[session_id] = req.username

        return {"otp_required": True, "session_id": session_id}
    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=500, detail="OTP failed")

@api_router.post("/verify-otp")
async def verify_otp(req: OTPVerifyRequest, session_id: str):
    if session_id not in pending_otp_sessions:
        raise HTTPException(status_code=400, detail="Invalid session")

    try:
        twilio = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        result = twilio.verify.v2.services(TWILIO_VERIFY_SID).verification_checks.create(
            to=ADMIN_PHONE, code=req.code
        )

        if result.status != "approved":
            raise HTTPException(status_code=400, detail="Invalid OTP")

        pending_otp_sessions.pop(session_id)
        token = create_access_token(
            {"sub": ADMIN_USER, "is_admin": True},
            ACCESS_TOKEN_EXPIRE_MINUTES
        )
        return {"token": token}

    except Exception as e:
        logger.error(e)
        raise HTTPException(status_code=500, detail="OTP verification failed")

@api_router.post("/admin/save")
async def save_student(student: StudentCreate, token: str):
    await get_current_admin(token)

    subjects = [{
        "semester": s.semester,
        "code": s.code,
        "grade": s.grade,
        "status": get_result_status(s.grade)
    } for s in student.subjects]

    await db.students.update_one(
        {"rollNo": student.rollNo},
        {"$set": {
            "rollNo": student.rollNo,
            "name": student.name,
            "dob": student.dob,
            "course": student.course,
            "subjects": subjects
        }},
        upsert=True
    )

    return {"message": "Saved successfully"}

# =====================
# EXCEL UPLOAD (MULTI SHEET + DOB FIX)
# =====================

@api_router.post("/admin/upload")
async def upload_excel(file: UploadFile = File(...), token: str = None):
    await get_current_admin(token)

    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        raw_headers = [str(c.value).strip().lower() for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            row_data = dict(zip(raw_headers, row))

            # 🔹 DOB FIX (dd/mm/yyyy handled)
            dob = row_data.get("dob", "")
            if isinstance(dob, (int, float)):
                dob = (datetime(1899, 12, 30) + timedelta(days=int(dob))).strftime("%Y-%m-%d")
            elif isinstance(dob, datetime):
                dob = dob.strftime("%Y-%m-%d")
            elif isinstance(dob, str):
                dob = dob.strip()
                try:
                    dob = datetime.strptime(dob, "%d/%m/%Y").strftime("%Y-%m-%d")
                except ValueError:
                    dob = dob
            else:
                dob = ""

            subjects = []
            for i in range(1, 26):
                sem = row_data.get(f"subjectsemester{i}")
                code = row_data.get(f"subjectcode{i}")
                grade = row_data.get(f"subjectgrade{i}")

                if sem and code and grade:
                    subjects.append({
                        "semester": str(sem),
                        "code": str(code),
                        "grade": str(grade),
                        "status": get_result_status(str(grade))
                    })

            await db.students.update_one(
                {"rollNo": str(row_data.get("rollno"))},
                {"$set": {
                    "rollNo": str(row_data.get("rollno")),
                    "name": str(row_data.get("name")),
                    "dob": dob,
                    "course": str(row_data.get("course")),
                    "subjects": subjects
                }},
                upsert=True
            )

    return {"message": "Excel uploaded successfully"}

@api_router.get("/student/result")
async def get_result(rollNo: str, dob: str):
    student = await db.students.find_one({"rollNo": rollNo, "dob": dob}, {"_id": 0})
    if not student:
        return {"message": "No result found"}

    return {
        "rollNo": student["rollNo"],
        "name": student["name"],
        "course": student["course"],
        "dob": student["dob"],
        "results": student["subjects"]
    }

@api_router.get("/verify-token")
async def verify_admin(token: str):
    return {"valid": bool(verify_token(token))}

# =====================
# FINAL
# =====================

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown():
    client.close()
